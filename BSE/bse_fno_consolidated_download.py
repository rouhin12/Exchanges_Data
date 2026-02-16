#!/usr/bin/env python3
"""
BSE F&O Daily Data Downloader (Consolidated Single File)

Downloads all F&O daily data and consolidates into a single file with NSE-style columns.
By default, downloads data for T-2 and T-1 (last 2 trading days).

Financial Year: April - March convention (FY2025_2026 = Apr 2025 - Mar 2026)

Column Structure (18 columns matching NSE format):
    Date, Index Futures (2 cols), Vol Futures (2 cols), Stock Futures (2 cols),
    Index Options (4 cols), Stock Options (4 cols), Total (3 cols)

Usage:
    # Download all 4 derivatives for last 2 days (default)
    python bse_fno_consolidated_download.py
    
    # Download specific derivatives
    python bse_fno_consolidated_download.py --derivatives index-futures equity-options
    
    # Download specific date range
    python bse_fno_consolidated_download.py --from-date 01/01/2026 --to-date 31/01/2026
    
    # Download specific months
    python bse_fno_consolidated_download.py --start-month 2026-01 --end-month 2026-02
"""

import sys
import argparse
from datetime import datetime, timedelta
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.edge.service import Service as EdgeService
import time
from openpyxl import Workbook, load_workbook
import logging

# Setup logging
logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)


# Wide table structure with all derivatives side-by-side (15 columns)
CONSOLIDATED_HEADERS = [
    "Date",
    # Index Options (4 columns)
    "Index Options - Number of Trades",
    "Index Options - Volume (Shares/Contracts)",
    "Index Options - Notional Turnover (₹ Cr)",
    "Index Options - Premium Turnover (₹ Cr)",
    # Stock Futures (3 columns)
    "Stock Futures - Number of Trades",
    "Stock Futures - Volume (Shares/Contracts)",
    "Stock Futures - Turnover (₹ Cr)",
    # Index Futures (3 columns)
    "Index Futures - Number of Trades",
    "Index Futures - Volume (Shares/Contracts)",
    "Index Futures - Turnover (₹ Cr)",
    # Stock Options (4 columns)
    "Stock Options - Number of Trades",
    "Stock Options - Volume (Shares/Contracts)",
    "Stock Options - Notional Turnover (₹ Cr)",
    "Stock Options - Premium Turnover (₹ Cr)"
]

# Mapping: (segment_name, product_type) -> (col_index_start, col_count)
# Column positions for each derivative (1-indexed, excluding Date)
COLUMN_MAPPING = {
    ("Index", "Futures"):         (1, 2),   # Cols 2-3
    ("Vol", "Futures"):           (3, 2),   # Cols 4-5
    ("Stock", "Futures"):         (5, 2),   # Cols 6-7
    ("Index", "Options"):         (7, 4),   # Cols 8-11
    ("Stock", "Options"):         (11, 4),  # Cols 12-15
    ("Total", ""):                (15, 3),  # Cols 16-18
}


def parse_date(value):
    """Parse dd-MMM-yyyy or dd/mm/yyyy format to datetime object."""
    try:
        # Try NSE format first (dd-MMM-yyyy)
        return datetime.strptime(value, "%d-%b-%Y")
    except ValueError:
        try:
            # Try alternative format (dd/mm/yyyy)
            return datetime.strptime(value, "%d/%m/%Y")
        except ValueError:
            raise ValueError(f"Invalid date format: {value}. Use dd-MMM-yyyy or dd/mm/yyyy")


def parse_month(value):
    """Parse MMM-yyyy format to list of (year, month) tuples for month range.
    
    Returns list of tuples for continuous months from given month to found month range.
    """
    try:
        dt = datetime.strptime(value, "%b-%Y")
        return [(dt.year, dt.month)]
    except ValueError:
        raise ValueError(f"Invalid month format: {value}. Use MMM-yyyy (e.g., Jan-2026)")


def parse_number(value):
    """Parse number from string with comma separators."""
    if not value or value == "":
        return 0
    try:
        return float(str(value).replace(",", ""))
    except:
        return 0


def format_number(value):
    """Format number with comma separators."""
    try:
        if isinstance(value, str):
            value = float(value.replace(",", ""))
        return f"{value:,.2f}"
    except:
        return ""


def get_last_two_days():
    """Get T-2 and T-1 (last 2 days)."""
    today = datetime.now()
    t_minus_1 = today - timedelta(days=1)
    t_minus_2 = today - timedelta(days=2)
    return t_minus_2, t_minus_1


def financial_year_start(date_obj):
    """Get the start year of the financial year (Apr-Mar convention)."""
    if date_obj.month < 4:
        return date_obj.year - 1
    return date_obj.year


def split_date_range(from_date, to_date, max_days=365):
    """Split date range into chunks of max_days."""
    chunks = []
    current_from = from_date
    
    while current_from <= to_date:
        current_to = current_from + timedelta(days=max_days - 1)
        if current_to > to_date:
            current_to = to_date
        
        chunks.append((current_from, current_to))
        current_from = current_to + timedelta(days=1)
    
    return chunks


def get_driver():
    """Initialize and return a Selenium webdriver."""
    options = webdriver.EdgeOptions()
    options.add_argument("--start-maximized")
    if "--headless" in sys.argv:
        options.add_argument("--headless=new")
    options.add_argument("--window-size=1400,900")
    
    return webdriver.Edge(options=options)


def download_segment_data(driver, segment_code_info, from_date, to_date):
    """Download F&O data from BSE for a specific segment and date range.
    
    Args:
        driver: Selenium webdriver instance
        segment_code_info: Tuple (segment_code, instrument_code, product_name)
                          E.g., ("ID", "FF", "Index Futures")
        from_date: datetime object
        to_date: datetime object
        
    Returns: Dict {date_str: [col_values]} or {} if error
    """
    url = "https://www.bseindia.com/markets/Derivatives/DeriReports/DeriHistoricalConsolidate.aspx"
    
    segment_code, instrument_code, product_name = segment_code_info
    
    try:
        driver.get(url)
        time.sleep(3)
        
        wait = WebDriverWait(driver, 10)
        
        # Select segment dropdown
        segment_select = wait.until(EC.presence_of_element_located((By.ID, "ContentPlaceHolder1_ddlsegment")))
        Select(segment_select).select_by_value(segment_code)
        logger.info(f"    Selected segment: {product_name}")
        time.sleep(2)
        
        # Select instrument dropdown  
        instr_select = driver.find_element(By.ID, "ContentPlaceHolder1_ddlIntrument")
        Select(instr_select).select_by_value(instrument_code)
        logger.info(f"    Selected instrument: {product_name} ({instrument_code})")
        time.sleep(1)
        
        # Format dates
        from_date_str = from_date.strftime("%d/%m/%Y")
        to_date_str = to_date.strftime("%d/%m/%Y")
        
        # Enter from date
        from_date_input = driver.find_element(By.ID, "ContentPlaceHolder1_txtDate")
        driver.execute_script(
            "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('change', { bubbles: true }));",
            from_date_input, from_date_str
        )
        time.sleep(1)
        
        # Enter to date
        to_date_input = driver.find_element(By.ID, "ContentPlaceHolder1_txtTodate")
        driver.execute_script(
            "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('change', { bubbles: true }));",
            to_date_input, to_date_str
        )
        time.sleep(1)
        
        # Click Go button
        go_button = driver.find_element(By.ID, "ContentPlaceHolder1_btnGo")
        driver.execute_script("arguments[0].click();", go_button)
        logger.info(f"    Waiting for data...")
        time.sleep(5)
        
        # Handle any alerts
        try:
            alert = WebDriverWait(driver, 2).until(EC.alert_is_present())
            alert.dismiss()
            time.sleep(1)
        except:
            pass
        
        # Extract table data
        tables = driver.find_elements(By.TAG_NAME, "table")
        if not tables:
            logger.warning(f"    No tables found")
            return {}
        
        # Find data table - look for the one with data rows
        data_table = None
        for table in tables:
            rows = table.find_elements(By.TAG_NAME, "tr")
            if len(rows) > 1:
                data_table = table
                break
        
        if not data_table:
            logger.warning(f"    No data table found")
            return {}
        
        # Extract rows
        rows = data_table.find_elements(By.TAG_NAME, "tr")
        date_data = {}
        
        for row in rows[1:]:  # Skip header row
            cells = row.find_elements(By.TAG_NAME, "td")
            if not cells or len(cells) < 2:
                continue
            
            # Extract cell values
            row_values = [cell.text.strip() for cell in cells]
            if not any(row_values):
                continue
            
            # First column is date (dd-mm-yyyy format)
            date_str = row_values[0]
            # Get all value columns
            values = row_values[1:]
            
            if date_str:
                # Convert date format from dd-mm-yyyy to dd-MMM-yyyy for consistency
                try:
                    date_obj = datetime.strptime(date_str, "%d-%m-%Y")
                    date_key = date_obj.strftime("%d-%b-%Y")
                    # BSE table structure: Trade Date | Number of Trades | Volume | Notional Turnover | Premium Turnover
                    # For futures: Trade Date | Number of Trades | Volume | Turnover (no premium)
                    # Store all available columns (3 for futures, 4 for options)
                    if "Options" in product_name:
                        date_data[date_key] = values[:4]  # Number of Trades, Volume, Notional, Premium
                    else:
                        date_data[date_key] = values[:3]  # Number of Trades, Volume, Turnover
                except:
                    pass
        
        logger.info(f"    Retrieved {len(date_data)} dates")
        return date_data
        
    except Exception as e:
        logger.error(f"    Error: {str(e)}")
        return {}


def merge_data_rows(date_str, index_fut, vol_fut, stock_fut, index_opt, stock_opt, total):
    """Merge data from 6 sources into single 18-column row."""
    row = [date_str]
    
    # Index Futures (2 cols): contracts, turnover
    row.extend(index_fut[:2] if index_fut and len(index_fut) >= 2 else ["", ""])
    
    # Vol Futures (2 cols)
    row.extend(vol_fut[:2] if vol_fut and len(vol_fut) >= 2 else ["", ""])
    
    # Stock Futures (2 cols)
    row.extend(stock_fut[:2] if stock_fut and len(stock_fut) >= 2 else ["", ""])
    
    # Index Options (4 cols): contracts, notional, premium, pcr
    row.extend(index_opt[:4] if index_opt and len(index_opt) >= 4 else ["", "", "", ""])
    
    # Stock Options (4 cols)
    row.extend(stock_opt[:4] if stock_opt and len(stock_opt) >= 4 else ["", "", "", ""])
    
    # Total (3 cols): contracts, turnover, pcr
    row.extend(total[:3] if total and len(total) >= 3 else ["", "", ""])
    
    return row


def save_consolidated(output_dir, consolidated_data):
    """Save consolidated data to a single workbook and single sheet.

    Args:
        output_dir: Output directory path
        consolidated_data: Dict {(year, month): [rows_with_18_cols]}
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = output_dir / "bse_fno_consolidated.xlsx"
    sheet_name = "All"

    # Flatten all rows into one list
    all_rows = []
    for rows in consolidated_data.values():
        all_rows.extend(rows)

    # Build map of existing rows by date to avoid duplicates
    existing_by_date = {}
    if filename.exists():
        try:
            
            wb = load_workbook(filename)
        except Exception:
            wb = Workbook()
            wb.remove(wb.active)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(CONSOLIDATED_HEADERS)

    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if not row or not row[0]:
            continue
        existing_by_date[str(row[0]).strip()] = True

    # Sort new rows by date and append only new dates
    sorted_rows = sorted(all_rows, key=lambda r: datetime.strptime(r[0], "%d-%b-%Y") if r[0] else datetime.min)
    appended = 0
    for row in sorted_rows:
        if len(row) < 15:
            row = row + [""] * (15 - len(row))
        elif len(row) > 15:
            row = row[:15]

        date_key = str(row[0]).strip() if row and row[0] else ""
        if date_key and date_key in existing_by_date:
            continue

        ws.append(row)
        if date_key:
            existing_by_date[date_key] = True
        appended += 1

    wb.save(filename)
    logger.info(f"  Saved: {filename.name} (sheet: {sheet_name}, +{appended} rows)")


def main():
    """Main entry point."""
    global logger
    
    parser = argparse.ArgumentParser(
        description="Download consolidated BSE F&O data (18-column NSE-style format)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python bse_fno_consolidated_download.py --start-month Jan-2026 --end-month Feb-2026
  python bse_fno_consolidated_download.py --start-date 01-Jan-2026 --end-date 31-Jan-2026
  python bse_fno_consolidated_download.py --year 2026
        """
    )
    
    date_group = parser.add_mutually_exclusive_group(required=True)
    date_group.add_argument("--date", type=str, default=None,
                          help="Single date (dd-MMM-yyyy)")
    date_group.add_argument("--start-date", "--sd", type=str, default=None,
                          help="Start date (dd-MMM-yyyy)")
    date_group.add_argument("--start-month", "--sm", type=str, default=None,
                          help="Start month (MMM-yyyy)")
    date_group.add_argument("--year", "-y", type=str, default=None,
                          help="Year (yyyy)")
    
    parser.add_argument("--end-date", "--ed", type=str, default=None,
                       help="End date (dd-MMM-yyyy)")
    parser.add_argument("--end-month", "--em", type=str, default=None,
                       help="End month (MMM-yyyy)")
    default_output = Path(__file__).resolve().parent / "FnO"
    parser.add_argument("--output-dir", "-o", type=str, default=str(default_output),
                       help="Output directory (default: BSE/FnO)")
    
    args = parser.parse_args()
    
    # Setup logging (use script directory as base)
    script_dir = Path(__file__).parent.parent.parent  # BSE/FnO -> BSE -> Exchanges Data
    log_dir = script_dir / "logs"
    log_dir.mkdir(exist_ok=True)
    log_file = log_dir / f"bse_fno_download_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.log"
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s | %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    logger = logging.getLogger(__name__)
    logger.info("=" * 80)
    logger.info("BSE F&O CONSOLIDATED DOWNLOADER (18-Column NSE-Style Format)")
    logger.info("=" * 80)
    
    # Determine date range
    if args.date:
        from_date = parse_date(args.date)
        to_date = from_date
    elif args.start_date:
        from_date = parse_date(args.start_date)
        to_date = parse_date(args.end_date) if args.end_date else from_date
    elif args.start_month:
        months_list = parse_month(args.start_month)
        from_date = datetime(months_list[0][0], months_list[0][1], 1)
        
        if args.end_month:
            end_months = parse_month(args.end_month)
            last_month = end_months[-1]
            to_date = datetime(last_month[0], last_month[1], 1) + timedelta(days=32)
            to_date = to_date.replace(day=1) - timedelta(days=1)
        else:
            to_date = datetime(months_list[-1][0], months_list[-1][1], 1) + timedelta(days=32)
            to_date = to_date.replace(day=1) - timedelta(days=1)
    elif args.year:
        year = int(args.year)
        from_date = datetime(year, 1, 1)
        to_date = datetime(year, 12, 31)
    
    logger.info(f"Date range: {from_date.strftime('%d-%b-%Y')} to {to_date.strftime('%d-%b-%Y')}")
    
    # Initialize browser
    options = EdgeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    
    driver = None
    try:
        driver = webdriver.Edge(options=options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        
        # Download all 4 segments and merge
        consolidated_data = {}  # Dict {(year, month): [rows]}
        
        # Download all 4 derivatives
        # Segment info: (segment_code, instrument_code, product_name)
        # Codes from BSE: Index Derivative (ID), Equity Derivative (ED)
        # Instruments: IF=Index Futures, IO=Index Options, SF=Stock Futures, SO=Stock Options
        segments = [
            ("ID", "IO", "Index Options"),
            ("ED", "SF", "Stock Futures"),
            ("ID", "IF", "Index Futures"),
            ("ED", "SO", "Stock Options")
        ]
        
        # Initialize date data structure to hold all segment data per date
        date_data = {}  # Dict {date_str: [18-col-row]}
        
        # Split date range into 365-day chunks (BSE limitation)
        date_chunks = split_date_range(from_date, to_date, max_days=365)
        total_days_span = (to_date - from_date).days + 1
        
        if len(date_chunks) > 1:
            logger.info(f"\nDate range spans {total_days_span} days - splitting into {len(date_chunks)} chunks of max 365 days")
            for i, (chunk_from, chunk_to) in enumerate(date_chunks, 1):
                logger.info(f"  Chunk {i}: {chunk_from.strftime('%d-%b-%Y')} to {chunk_to.strftime('%d-%b-%Y')}")
        
        logger.info(f"\nDownloading {len(segments)} segments...")
        for seg_info in segments:
            segment_code, instrument_code, seg_name = seg_info
            logger.info(f"\n[>] {seg_name}...")
            
            try:
                # Download data in chunks and merge
                segment_data = {}
                
                for chunk_idx, (chunk_from, chunk_to) in enumerate(date_chunks, 1):
                    if len(date_chunks) > 1:
                        logger.info(f"    Downloading chunk {chunk_idx}/{len(date_chunks)}...")
                    
                    chunk_data = download_segment_data(driver, seg_info, chunk_from, chunk_to)
                    
                    if chunk_data:
                        segment_data.update(chunk_data)
                        if len(date_chunks) > 1:
                            logger.info(f"    Chunk {chunk_idx}: {len(chunk_data)} dates retrieved")
                    
                    # Add delay between chunks to avoid rate limiting
                    if chunk_idx < len(date_chunks):
                        time.sleep(2)
                
                if not segment_data:
                    logger.warning(f"    No data returned for {seg_name}")
                    continue
                
                logger.info(f"    Total retrieved: {len(segment_data)} dates")
                
                # Merge segment data into consolidated structure
                for date_str, values in segment_data.items():
                    if date_str not in date_data:
                        # Initialize 15-column row: [date, "", ..., ""]
                        date_data[date_str] = [date_str] + [""] * 14
                    
                    # Place each derivative's data in the correct columns
                    if "Index Options" in seg_name:
                        # Cols 1-4: Index Options (Number of Trades, Volume, Notional, Premium)
                        date_data[date_str][1:5] = values[:4]
                    elif "Stock Futures" in seg_name:
                        # Cols 5-7: Stock Futures (Number of Trades, Volume, Turnover)
                        date_data[date_str][5:8] = values[:3]
                    elif "Index Futures" in seg_name:
                        # Cols 8-10: Index Futures (Number of Trades, Volume, Turnover)
                        date_data[date_str][8:11] = values[:3]
                    elif "Stock Options" in seg_name:
                        # Cols 11-14: Stock Options (Number of Trades, Volume, Notional, Premium)
                        date_data[date_str][11:15] = values[:4]
                
                logger.info(f"    [OK] {seg_name} merged")
                
            except Exception as e:
                logger.error(f"    [X] Error downloading {seg_name}: {str(e)}")
                continue
        
        # No need to calculate totals for simplified format
        logger.info(f"\n[>] Processing {len(date_data)} dates...")
        
        # Group by financial year and month
        logger.info(f"\nGrouping by financial year...")
        for date_str, row in date_data.items():
            try:
                date_obj = datetime.strptime(date_str, "%d-%b-%Y")
                fy_start = financial_year_start(date_obj)
                month_key = (date_obj.year, date_obj.month)
                
                if month_key not in consolidated_data:
                    consolidated_data[month_key] = []
                
                consolidated_data[month_key].append(row)
            except Exception as e:
                logger.warning(f"    Error processing date {date_str}: {str(e)}")
        
        # Save consolidated data
        logger.info(f"\nSaving consolidated data...")
        save_consolidated(args.output_dir, consolidated_data)
        
        total_rows = sum(len(rows) for rows in consolidated_data.values())
        logger.info(f"\n[OK] Download Complete! Total days downloaded: {total_rows}")
        logger.info("=" * 80)
        
    except Exception as e:
        logger.error(f"[X] Fatal error: {str(e)}")
        raise
    finally:
        if driver:
            driver.quit()


if __name__ == "__main__":
    main()
