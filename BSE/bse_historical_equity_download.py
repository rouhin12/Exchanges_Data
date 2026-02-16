import argparse
import datetime as dt
import os
import time

from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

URL = "https://www.bseindia.com/markets/Equity/EQReports/Historical_EquitySegment.aspx"


def parse_month(value):
    try:
        return dt.datetime.strptime(value, "%Y-%m").date().replace(day=1)
    except ValueError as exc:
        raise ValueError("Month must be in YYYY-MM format.") from exc


def iter_months(start_date, end_date):
    current = start_date.replace(day=1)
    end = end_date.replace(day=1)
    while current <= end:
        yield current
        year = current.year + (current.month // 12)
        month = 1 if current.month == 12 else current.month + 1
        current = current.replace(year=year, month=month)


def build_driver(download_dir, headless):
    options = webdriver.EdgeOptions()
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
    }
    options.add_experimental_option("prefs", prefs)
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--window-size=1400,900")
    return webdriver.Edge(options=options)


def find_link_contains(driver, text):
    xpath = f"//a[contains(normalize-space(.), '{text}')]"
    for element in driver.find_elements(By.XPATH, xpath):
        if element.is_displayed():
            return element
    return None


def safe_click(driver, element):
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
    time.sleep(0.3)
    try:
        element.click()
        return
    except Exception:
        pass
    driver.execute_script("arguments[0].click();", element)


def wait_for_ready_state(driver, timeout=20):
    WebDriverWait(driver, timeout).until(lambda d: d.execute_script("return document.readyState") == "complete")


def debug_api_resources(driver):
    try:
        resources = driver.execute_script(
            "return performance.getEntriesByType('resource').map(r => r.name);"
        )
    except Exception as exc:
        print(f"Debug: unable to read performance entries: {exc}")
        return

    if not resources:
        print("Debug: no resource entries found.")
        return

    filtered = []
    for url in resources:
        lower = url.lower()
        if any(token in lower for token in ("api", "json", "ashx", "svc", "handler")):
            filtered.append(url)

    unique = list(dict.fromkeys(filtered or resources))
    print("Debug: resource URLs (filtered):")
    for url in unique[:200]:
        print(f"  {url}")


def financial_year_start(date_value):
    return date_value.year - 1 if date_value.month <= 3 else date_value.year


def click_year(driver, year):
    candidates = [f"{year}-{year + 1}", str(year)]
    for candidate in candidates:
        link = find_link_contains(driver, candidate)
        if link:
            safe_click(driver, link)
            return True
    return False


def click_month(driver, date_value):
    month_name = date_value.strftime("%B")
    month_abbr = date_value.strftime("%b")
    year_short = date_value.strftime("%y")
    candidates = [
        f"{month_name} {date_value.year}",
        f"{month_abbr}-{year_short}",
        f"{month_abbr} {year_short}",
        month_name,
        month_abbr,
    ]
    for candidate in candidates:
        link = find_link_contains(driver, candidate)
        if link:
            safe_click(driver, link)
            return True
    return False


def wait_for_daily_table(driver, timeout=20):
    wait = WebDriverWait(driver, timeout)
    return wait.until(EC.presence_of_element_located((By.ID, "ContentPlaceHolder1_grddaily")))


def read_daily_table(table):
    header_cells = table.find_elements(By.CSS_SELECTOR, "tr th")
    header = [cell.text.strip() for cell in header_cells]

    rows = []
    for tr in table.find_elements(By.CSS_SELECTOR, "tr")[1:]:
        cells = tr.find_elements(By.CSS_SELECTOR, "td")
        if not cells:
            continue
        row = [cell.text.strip() for cell in cells]
        if any(row):
            rows.append(row)
    return header, rows


def parse_date_from_cell(date_str):
    date_str = str(date_str).strip()
    for fmt in ["%d-%b-%Y", "%d-%B-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d %b %Y", "%d %B %Y"]:
        try:
            return dt.datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def save_month_to_workbook(output_dir, date_value, header, rows):
    filename = os.path.join(output_dir, "bse_daily.xlsx")
    sheet_name = "All"

    if os.path.exists(filename):
        workbook = load_workbook(filename)
    else:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(header)

    existing_dates = set()
    for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if not row or row[0] is None:
            continue
        parsed = parse_date_from_cell(row[0])
        if parsed:
            existing_dates.add(parsed)

    for row in rows:
        row_date = parse_date_from_cell(row[0]) if row else None
        if row_date and row_date in existing_dates:
            continue
        worksheet.append(row)
        if row_date:
            existing_dates.add(row_date)

    workbook.save(filename)
    return filename


def main():
    parser = argparse.ArgumentParser(description="Download BSE daily data month-by-month.")
    parser.add_argument("--start-month", default="2020-04", help="Start month in YYYY-MM format.")
    parser.add_argument(
        "--end-month",
        default=dt.date.today().strftime("%Y-%m"),
        help="End month in YYYY-MM format.",
    )
    default_output = os.path.dirname(os.path.abspath(__file__))
    parser.add_argument("--output-dir", default=default_output, help="Output folder for CSV files.")
    parser.add_argument("--headless", action="store_true", help="Run Chrome in headless mode.")
    parser.add_argument("--debug-api", action="store_true", help="Print network resource URLs and exit.")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    start_date = parse_month(args.start_month)
    end_date = parse_month(args.end_month)

    driver = build_driver(args.output_dir, args.headless)
    try:
        if args.debug_api:
            driver.get(URL)
            wait_for_ready_state(driver)
            time.sleep(1)
            debug_api_resources(driver)
            return
        for date_value in iter_months(start_date, end_date):
            print(f"Processing {date_value.strftime('%Y-%m')}...")
            try:
                driver.get(URL)
                wait_for_ready_state(driver)

                fy_start = financial_year_start(date_value)
                if not click_year(driver, fy_start):
                    print(f"Could not select financial year {fy_start}-{fy_start + 1}. Skipping.")
                    continue

                time.sleep(1)

                if not click_month(driver, date_value):
                    print(f"Could not select month {date_value.strftime('%b %Y')}. Skipping.")
                    continue
                wait_for_ready_state(driver)
                time.sleep(1)

                table = wait_for_daily_table(driver)
                header, rows = read_daily_table(table)
                if not header or not rows:
                    print(f"No daily rows found for {date_value.strftime('%Y-%m')}.")
                    continue

                saved_path = save_month_to_workbook(args.output_dir, date_value, header, rows)
                print(f"Saved: {saved_path} ({date_value.strftime('%Y-%m')})")
            except Exception as exc:
                print(f"Error on {date_value.strftime('%Y-%m')}: {exc}")
                driver.quit()
                driver = build_driver(args.output_dir, args.headless)
                continue
    finally:
        driver.quit()


if __name__ == "__main__":
    main()
