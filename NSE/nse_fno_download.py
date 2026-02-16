#!/usr/bin/env python3
"""
NSE F&O Segment Daily Data Downloader (Selenium-based web scraper)
Downloads daily F&O data by navigating through the year/month links on NSE website.
Uses Notional Turnover view as specified.

Usage:
    # Download current month (default)
    python nse_fno_download.py
    
    # Download specific date range
    python nse_fno_download.py --from-date 01/01/2026 --to-date 31/01/2026
    
    # Download specific months
    python nse_fno_download.py --start-month Jan-2026 --end-month Feb-2026
    
    # Download financial years
    python nse_fno_download.py --start-year 2024 --end-year 2025
"""
import os
import sys
import argparse
import datetime as dt
import time
import re
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from openpyxl import Workbook, load_workbook


TARGET_URL = "https://www.nseindia.com/market-data/business-growth-fo-segment"


def financial_year_start(date_obj):
    """Return the starting year of the financial year (April-March)."""
    if date_obj.month >= 4:
        return date_obj.year
    else:
        return date_obj.year - 1


def get_last_two_days():
    """Get T-2 and T-1 business days (approximation)."""
    today = dt.date.today()
    t_minus_1 = today - timedelta(days=1)
    t_minus_2 = today - timedelta(days=2)
    return t_minus_2, t_minus_1


def setup_driver():
    """Initialize Edge WebDriver with appropriate options."""
    options = EdgeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    
    driver = webdriver.Edge(options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


def save_month_to_workbook(output_dir, date_value, header, rows):
    """Save data rows to a single workbook and single sheet, appending new dates."""
    filename = os.path.join(output_dir, "nse_fno_consolidated.xlsx")
    sheet_name = "All"

    if os.path.exists(filename):
        try:
            workbook = load_workbook(filename)
        except Exception as e:
            # File is corrupted - delete and recreate
            print(f"  [!] Warning: Corrupted file detected ({e}), recreating...")
            os.remove(filename)
            workbook = Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)
    else:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(header)

    # Build a map of existing rows by date to prevent duplicates
    existing_by_date = {}
    extra_rows = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        row_date = parse_date_from_cell(str(row[0])) if row[0] else None
        if row_date:
            existing_by_date[row_date] = list(row)
        else:
            extra_rows.append(list(row))
    
    # Append all data rows - pad rows to match header count if needed
    expected_cols = len(header)
    for row in rows:
        # Ensure row has same number of columns as header
        if len(row) < expected_cols:
            row = row + [""] * (expected_cols - len(row))
        elif len(row) > expected_cols:
            row = row[:expected_cols]

        row_date = parse_date_from_cell(str(row[0])) if row and row[0] else None
        if row_date:
            row[0] = row_date.strftime("%d-%b-%Y")
            existing_by_date[row_date] = row
        else:
            extra_rows.append(row)

    # Rewrite sheet in chronological order (oldest to newest)
    sorted_rows = [row for _, row in sorted(existing_by_date.items(), key=lambda item: item[0])]
    sorted_rows.extend(extra_rows)
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)
    for row in sorted_rows:
        worksheet.append(row)

    workbook.save(filename)
    return filename


def parse_date(date_str):
    """Parse date from dd/mm/yyyy format to datetime object."""
    try:
        return datetime.strptime(date_str, "%d/%m/%Y")
    except ValueError:
        raise ValueError(f"Invalid date format: {date_str}. Use dd/mm/yyyy")


def parse_date_from_cell(date_str):
    """Parse date from table cell (handles various formats)."""
    date_str = date_str.strip()
    
    # Try common date formats
    for fmt in ["%d-%b-%Y", "%d-%b-%y", "%d-%B-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d %b %Y", "%d %b %y"]:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    
    return None


def clean_numeric_value(value_str):
    """Remove commas and convert to numeric string, or return empty."""
    value_str = value_str.strip()
    if value_str == "-" or value_str == "":
        return ""
    # Remove commas
    return value_str.replace(",", "")


def scrape_daily_table(driver):
    """Scrape the daily data table from the current page."""
    try:
        # Wait for the table container - use shorter timeout since table should already be loaded
        wait = WebDriverWait(driver, 5)
        
        # Try to find the monthly detail table first (appears after clicking month)
        try:
            table_container = wait.until(
                EC.presence_of_element_located((By.ID, "tbgDeriv_month_detail"))
            )
            table = table_container.find_element(By.CSS_SELECTOR, "table.common_table")
        except:
            # Fallback to general table
            table = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.common_table")))
        
        # Extract headers from the table
        # Hardcoded headers based on NSE F&O table structure with Notional Turnover enabled
        headers = [
            "Date",
            "Index Futures - No. Of Contracts",
            "Index Futures - Turnover (₹ Crores)",
            "Vol Futures - No. Of Contracts",
            "Vol Futures - Turnover (₹ Crores)",
            "Stock Futures - No. Of Contracts",
            "Stock Futures - Turnover (₹ Crores)",
            "Index Options - No. Of Contracts",
            "Index Options - Notional Turnover (₹ Crores)",
            "Index Options - Premium Turnover (₹ Crores)",
            "Index Options - Put Call Ratio",
            "Stock Options - No. Of Contracts",
            "Stock Options - Notional Turnover (₹ Crores)",
            "Stock Options - Premium Turnover (₹ Crores)",
            "Stock Options - Put Call Ratio",
            "Total - No. Of Contracts",
            "Total - Turnover (₹ Crores)",
            "Total - Put Call Ratio"
        ]
        
        # Extract data rows
        rows = []
        tbody = table.find_element(By.TAG_NAME, "tbody")
        tr_elements = tbody.find_elements(By.TAG_NAME, "tr")
        
        for tr in tr_elements:
            # Get ALL td cells including hidden ones
            cells = driver.execute_script("""
                return arguments[0].querySelectorAll('td');
            """, tr)
            
            row_data = []
            
            for cell in cells:
                # Get text content directly from JavaScript to include hidden cells
                text = driver.execute_script("""
                    var cell = arguments[0];
                    // Try to get link text first
                    var link = cell.querySelector('a');
                    if (link) return link.textContent.trim();
                    return cell.textContent.trim();
                """, cell)
                
                # Keep original value (don't clean yet to preserve dates)
                row_data.append(text if text else "")
            
            if row_data and any(row_data):  # Only add non-empty rows
                rows.append(row_data)
        
        return headers, rows
    
    except Exception as e:
        print(f"[X] Error scraping table: {e}")
        import traceback
        traceback.print_exc()
        return [], []


def download_daily_data_for_month(driver, year_text, month_text, output_dir, date_filter=None):
    """Navigate to a specific month and download daily data.
    
    Args:
        year_text: Full year text like '2023-2024'
        month_text: Month text like 'Feb-2024'
        date_filter: Optional tuple of (from_date, to_date) datetime objects to filter data
    """
    try:
        wait = WebDriverWait(driver, 12)  # Balanced timeout - table usually loads quickly
        
        # First, try to reset year view by finding collapsed year links via index
        # This handles the case where previous navigation cleared link text
        year_links = driver.find_elements(By.CSS_SELECTOR, "a.year_link")
        fy_start_year = int(year_text.split("-")[0])
        
        year_link = None
        year_link_index = None
        
        # Try to find by exact text match first
        for idx, link in enumerate(year_links):
            try:
                if link.text.strip() == year_text:
                    year_link = link
                    year_link_index = idx
                    break
            except:
                pass
        
        # If text match failed, try to deduce index from year number
        # Years are in descending order (newest first)
        # 2025-2026 is at index 0, 2024-2025 at index 1, etc.
        if not year_link:
            try:
                current_year = int(year_text.split("-")[0])
                # All year links in reverse order
                all_year_texts = [f"{2025-i}-{2026-i}" for i in range(11)]  # 2025-2016
                if year_text in all_year_texts:
                    year_link_index = all_year_texts.index(year_text)
                    if year_link_index < len(year_links):
                        year_link = year_links[year_link_index]
            except:
                pass
        
        if not year_link:
            return 0
        
        # Click year to reveal months
        try:
            driver.execute_script("arguments[0].click();", year_link)
            time.sleep(0.5)  # Reduced from 1s
        except StaleElementReferenceException:
            # Re-find by index and click
            year_links = driver.find_elements(By.CSS_SELECTOR, "a.year_link")
            if year_link_index is not None and year_link_index < len(year_links):
                driver.execute_script("arguments[0].click();", year_links[year_link_index])
            time.sleep(0.5)  # Reduced from 1s
        
        # Find and click month link using exact text match
        month_link = None
        month_links = driver.find_elements(By.CSS_SELECTOR, "a.month_link")
        for link in month_links:
            try:
                if link.text.strip() == month_text:
                    month_link = link
                    break
            except StaleElementReferenceException:
                continue
        
        if not month_link:
            return 0
        
        # Click month to show daily data
        try:
            driver.execute_script("arguments[0].click();", month_link)
            time.sleep(1)  # Reduced from 1.5s
        except StaleElementReferenceException:
            month_links = driver.find_elements(By.CSS_SELECTOR, "a.month_link")
            for link in month_links:
                if link.text.strip() == month_text:
                    driver.execute_script("arguments[0].click();", link)
                    break
            time.sleep(1)  # Reduced from 1.5s
        
        # Wait for the monthly detail table to appear
        try:
            wait.until(EC.presence_of_element_located((By.ID, "tbgDeriv_month_detail")))
        except TimeoutException:
            pass
        
        # Click notional button every month (state doesn't persist across navigation)
        try:
            notional_btn = driver.find_element(By.ID, "showNotionalTurnover")
            driver.execute_script("arguments[0].click();", notional_btn)
            time.sleep(0.5)  # Wait for the DOM to update after button click
            
            # Force show all notional columns
            driver.execute_script("""
                var selectors = [
                    '#notionalHide_Index_Daily', '#notionalHide_Stock_Daily',
                    '#total_turnover', '[id*="notional"]', '[id*="Notional"]'
                ];
                selectors.forEach(function(selector) {
                    try {
                        document.querySelectorAll(selector).forEach(function(el) {
                            el.style.display = '';
                            el.style.visibility = 'visible';
                            el.style.width = 'auto';
                        });
                    } catch(e) {}
                });
            """)
            
            time.sleep(0.5)  # Additional wait for values to populate
        except:
            pass
        
        # Scrape the daily table (notional columns enabled on first month, persists for all others)
        headers, rows = scrape_daily_table(driver)
        
        if not rows:
            return 0
        
        # Save each day's data
        saved_count = 0
        for row in rows:
            if not row or len(row) == 0:
                continue
            
            # First column should be the date
            date_str = row[0]
            date_value = parse_date_from_cell(date_str)
            
            if not date_value:
                continue
            
            # Apply date filter if specified
            if date_filter:
                from_date, to_date = date_filter
                if date_value < from_date.date() or date_value > to_date.date():
                    continue
            
            # Save this day's data
            save_month_to_workbook(output_dir, date_value, headers, [row])
            saved_count += 1
        
        return saved_count
    
    except Exception as e:
        print(f"  [X] Error: {e}")
        import traceback
        traceback.print_exc()
        return 0


def main():
    parser = argparse.ArgumentParser(description="Download NSE F&O segment daily data via web scraping.")
    parser.add_argument("--from-date", help="From date (dd/mm/yyyy format). If not provided, defaults to current month.")
    parser.add_argument("--to-date", help="To date (dd/mm/yyyy format). If not provided, defaults to current month.")
    parser.add_argument("--start-year", type=int, help="Start financial year (e.g., 2025 for FY2025-2026).")
    parser.add_argument("--end-year", type=int, help="End financial year.")
    parser.add_argument("--start-month", help="Start month in MMM-YYYY format (e.g., Feb-2026).")
    parser.add_argument("--end-month", help="End month in MMM-YYYY format (e.g., Feb-2026).")
    default_output = os.path.join(os.path.dirname(os.path.abspath(__file__)), "FnO")
    parser.add_argument("--output-dir", default=default_output, help="Output folder for Excel files.")
    parser.add_argument("--headless", action="store_true", help="Run browser in headless mode.")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    
    # Determine what to download
    months_to_download = []
    date_filter = None  # (from_date, to_date) for filtering daily data
    
    # Priority 1: Date range (--from-date and --to-date)
    if args.from_date or args.to_date:
        try:
            from_date = parse_date(args.from_date) if args.from_date else datetime.now().replace(day=1)
            to_date = parse_date(args.to_date) if args.to_date else datetime.now()
            
            if from_date > to_date:
                print("Error: from-date cannot be after to-date")
                return
            
            date_filter = (from_date, to_date)
            
            # Generate list of months to download
            current = from_date
            seen_months = set()
            while current <= to_date:
                month_abbr = current.strftime("%b")
                month_year = current.year
                month_key = (month_abbr, month_year)
                if month_key not in seen_months:
                    months_to_download.append(month_key)
                    seen_months.add(month_key)
                # Move to next month
                if current.month == 12:
                    current = current.replace(year=current.year + 1, month=1)
                else:
                    current = current.replace(month=current.month + 1)
            
            print(f"\n{'='*70}")
            print(f"NSE F&O Daily Data Downloader (Web Scraper)")
            print(f"Date Range: {from_date.strftime('%d/%m/%Y')} to {to_date.strftime('%d/%m/%Y')}")
            print(f"Output: {args.output_dir}")
            print(f"{'='*70}\n")
        except ValueError as e:
            print(f"Error: {e}")
            return
    
    # Priority 2: Specific months (--start-month and --end-month)
    elif args.start_month and args.end_month:
        # Parse month range
        # Format: Feb-2026
        start_parts = args.start_month.split("-")
        end_parts = args.end_month.split("-")
        
        if len(start_parts) == 2 and len(end_parts) == 2:
            start_month_abbr = start_parts[0]
            start_year = int(start_parts[1])
            end_month_abbr = end_parts[0]
            end_year = int(end_parts[1])
            
            # Generate list of months between start and end
            all_months = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
            
            # Iterate from start month/year to end month/year
            current_year = start_year
            current_month_idx = all_months.index(start_month_abbr) if start_month_abbr in all_months else 0
            end_month_idx = all_months.index(end_month_abbr) if end_month_abbr in all_months else 0
            
            # Add all months in range
            added = set()
            while True:
                month_abbr = all_months[current_month_idx]
                month_key = (month_abbr, current_year)
                if month_key not in added:
                    months_to_download.append(month_key)
                    added.add(month_key)
                
                # Stop if we reached the end month and year
                if current_month_idx == end_month_idx and current_year == end_year:
                    break
                
                # Increment year when moving from December to January (calendar year boundary)
                if all_months[current_month_idx] == "Dec":
                    current_year += 1
                
                # Move to next month
                current_month_idx = (current_month_idx + 1) % 12
        
        print(f"\n{'='*70}")
        print(f"NSE F&O Daily Data Downloader (Web Scraper)")
        print(f"Months: {args.start_month} to {args.end_month}")
        print(f"Output: {args.output_dir}")
        print(f"{'='*70}\n")
    
    # Priority 3: Financial year range (--start-year and --end-year)
    elif args.start_year and args.end_year:
        start_year = args.start_year
        end_year = args.end_year
        
        print(f"\n{'='*70}")
        print(f"NSE F&O Daily Data Downloader (Web Scraper)")
        print(f"Financial Years: {start_year}-{start_year+1} to {end_year}-{end_year+1}")
        print(f"Output: {args.output_dir}")
        print(f"{'='*70}\n")
        
        # Generate months for each FY
        all_months = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
        for fy in range(start_year, end_year + 1):
            for month_abbr in all_months:
                if month_abbr in ["Jan", "Feb", "Mar"]:
                    month_year = fy + 1
                else:
                    month_year = fy
                months_to_download.append((month_abbr, month_year))
    
    # Default: Current month
    else:
        today = datetime.now()
        current_month_abbr = today.strftime("%b")
        current_year = today.year
        months_to_download.append((current_month_abbr, current_year))
        
        print(f"\n{'='*70}")
        print(f"NSE F&O Daily Data Downloader (Web Scraper)")
        print(f"Default: Current month ({current_month_abbr}-{current_year})")
        print(f"Output: {args.output_dir}")
        print(f"{'='*70}\n")

    # Setup Selenium driver
    driver = setup_driver()
    
    try:
        # Navigate to NSE F&O page
        print("Opening NSE F&O page...")
        driver.get(TARGET_URL)
        time.sleep(5)  # Let page load completely
        
        # Wait for the main table to be visible
        try:
            wait = WebDriverWait(driver, 10)
            wait.until(EC.presence_of_element_located((By.ID, "tbgDeriv_detail_table")))
        except TimeoutException:
            print("⚠ Main table not found, proceeding anyway...")
        
        print("Note: Notional turnover will be enabled monthly during downloads\n")
        time.sleep(0.5)
        
        # Process each month
        total_days = 0
        for month_abbr, month_year in months_to_download:
            # Determine FY for this month
            if month_abbr in ["Jan", "Feb", "Mar"]:
                fy = month_year - 1
            else:
                fy = month_year
            
            year_text = f"{fy}-{fy+1}"
            month_text = f"{month_abbr}-{month_year}"
            
            print(f"{month_text}...", end=" ", flush=True)
            
            # Download daily data for this month
            count = download_daily_data_for_month(driver, year_text, month_text, args.output_dir, date_filter)
            
            if count > 0:
                print(f"[OK] {count} days")
                total_days += count
            else:
                print("[X] No data")
            
            # Minimal wait between months instead of full page reload
            time.sleep(0.5)
        
        print(f"\n{'='*70}")
        print(f"[OK] Download Complete! Total days downloaded: {total_days}")
        print(f"{'='*70}\n")
    
    finally:
        driver.quit()


if __name__ == "__main__":
    main()
