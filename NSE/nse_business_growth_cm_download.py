import argparse
import datetime as dt
import json
import os
import re
import time
import urllib.error
import urllib.request
from http.cookiejar import CookieJar

from openpyxl import Workbook, load_workbook

URL = "https://www.nseindia.com/market-data/business-growth-cm-segment"
API_URL = "https://www.nseindia.com/api/historicalOR/cm/tbg/daily?month={month}&year={year}"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": URL,
    "Origin": "https://www.nseindia.com",
    "Connection": "keep-alive",
    "DNT": "1",
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-origin",
    "sec-ch-ua": "\"Not.A/Brand\";v=8, \"Chromium\";v=120, \"Google Chrome\";v=120",
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": "\"Windows\"",
}


def parse_month(value):
    try:
        return dt.datetime.strptime(value, "%Y-%m").date().replace(day=1)
    except ValueError as exc:
        raise ValueError("Month must be in YYYY-MM format.") from exc


def iter_months(start_date, end_date):
    current = start_date.replace(day=1)
    end = end_date.replace(day=1)
    while current <= end:
        yield current
        year = current.year + (current.month // 12)
        month = 1 if current.month == 12 else current.month + 1
        current = current.replace(year=year, month=month)


def financial_year_start(date_value):
    return date_value.year - 1 if date_value.month <= 3 else date_value.year


def build_opener():
    cookie_jar = CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cookie_jar))
    return opener


def fetch_json(opener, url, retries=1):
    request = urllib.request.Request(url, headers=HEADERS)
    try:
        with opener.open(request, timeout=30) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        if exc.code in (401, 403) and retries > 0:
            prime_session(opener)
            return fetch_json(opener, url, retries=retries - 1)
        raise


def prime_session(opener):
    priming_headers = dict(HEADERS)
    priming_headers["Accept"] = "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
    priming_headers["Sec-Fetch-Dest"] = "document"
    priming_headers["Sec-Fetch-Mode"] = "navigate"
    priming_headers["Sec-Fetch-Site"] = "none"

    for url in ("https://www.nseindia.com", URL):
        request = urllib.request.Request(url, headers=priming_headers)
        with opener.open(request, timeout=30):
            time.sleep(0.5)


def normalize_key(key):
    return re.sub(r"[^a-z0-9]", "", key.lower())


def order_keys(keys):
    preferred = [
        "ftimestamp",
        "cdtnosofsecuritytrades",
        "cdtnosoftrades",
        "cdttradesqty",
        "cdttradesvalues",
        "type",
        "cdtdateorder",
        "date",
        "noofsecuritiestraded",
        "nooftrades",
        "tradedquantity",
        "tradedvalue",
    ]
    normalized = {normalize_key(key): key for key in keys}
    ordered = []
    for pref in preferred:
        if pref in normalized:
            ordered.append(normalized[pref])
    for key in keys:
        if key not in ordered:
            ordered.append(key)
    return ordered


def extract_rows(payload):
    if not isinstance(payload, dict):
        return []

    for key in ("data", "records", "rows"):
        value = payload.get(key)
        if isinstance(value, list):
            rows = []
            for item in value:
                if isinstance(item, dict) and "data" in item:
                    rows.append(item["data"])
                else:
                    rows.append(item)
            return rows
    return []


def parse_date_from_cell(date_str):
    date_str = str(date_str).strip()
    for fmt in ["%d-%b-%Y", "%d-%B-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d %b %Y", "%d %B %Y"]:
        try:
            return dt.datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def save_month_to_workbook(output_dir, date_value, header, rows):
    filename = os.path.join(output_dir, "nse_daily.xlsx")
    sheet_name = "All"

    if os.path.exists(filename):
        workbook = load_workbook(filename)
    else:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        existing_header = [cell.value for cell in worksheet[1]]
        if any(existing_header):
            header = existing_header
    else:
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(header)

    normalized_header = [normalize_key(str(col)) for col in header]
    date_col_index = None
    if "date" in normalized_header:
        date_col_index = normalized_header.index("date")
    elif "ftimestamp" in normalized_header:
        date_col_index = normalized_header.index("ftimestamp")

    existing_rows = []
    existing_dates = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        row_date_value = None
        if date_col_index is not None and date_col_index < len(row):
            row_date_value = row[date_col_index]
        elif len(row) > 0:
            row_date_value = row[0]
        parsed = parse_date_from_cell(row_date_value) if row_date_value is not None else None
        if parsed:
            existing_dates[parsed] = list(row)
        else:
            existing_rows.append(list(row))

    expected_cols = len(header)
    for row in rows:
        if row is None:
            continue
        if len(row) < expected_cols:
            row = row + [""] * (expected_cols - len(row))
        elif len(row) > expected_cols:
            row = row[:expected_cols]

        if date_col_index is not None and row[date_col_index] is not None:
            row_date = parse_date_from_cell(row[date_col_index])
        else:
            row_date = parse_date_from_cell(row[0]) if row else None
        if row_date:
            existing_dates[row_date] = row
        else:
            existing_rows.append(row)

    sorted_dates = sorted(existing_dates.items(), key=lambda item: item[0])
    ordered_rows = [row for _, row in sorted_dates] + existing_rows

    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)
    for row in ordered_rows:
        worksheet.append(row)

    workbook.save(filename)
    return filename


def main():
    parser = argparse.ArgumentParser(description="Download NSE CM segment daily data month-by-month.")
    parser.add_argument("--start-month", default="2020-04", help="Start month in YYYY-MM format.")
    parser.add_argument(
        "--end-month",
        default=dt.date.today().strftime("%Y-%m"),
        help="End month in YYYY-MM format.",
    )
    default_output = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Cash segment")
    parser.add_argument("--output-dir", default=default_output, help="Output folder for Excel files.")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    start_date = parse_month(args.start_month)
    end_date = parse_month(args.end_month)

    opener = build_opener()
    prime_session(opener)

    for date_value in iter_months(start_date, end_date):
        month_label = date_value.strftime("%b")
        year_label = date_value.strftime("%y")
        print(f"Processing {date_value.strftime('%Y-%m')}...")

        url = API_URL.format(month=month_label, year=year_label)
        try:
            payload = fetch_json(opener, url, retries=1)
        except urllib.error.HTTPError as exc:
            print(f"HTTP error for {date_value.strftime('%Y-%m')}: {exc}")
            continue
        except urllib.error.URLError as exc:
            print(f"Network error for {date_value.strftime('%Y-%m')}: {exc}")
            continue

        rows = extract_rows(payload)
        if not rows:
            print(f"No data returned for {date_value.strftime('%Y-%m')}")
            continue

        if isinstance(rows[0], dict):
            column_map = [
                ("F_TIMESTAMP", "Date"),
                ("CDT_NOS_OF_SECURITY_TRADES", "No. of securities traded"),
                ("CDT_NOS_OF_TRADES", "No of trades"),
                ("CDT_TRADES_QTY", "Traded quantity (in Lakhs)"),
                ("CDT_TRADES_VALUES", "Traded Value (₹ Crores)"),
            ]
            header = [label for _, label in column_map]
            data_rows = [[row.get(key, "") for key, _ in column_map] for row in rows]
        else:
            header = ["data"]
            data_rows = [[value] for value in rows]

        saved_path = save_month_to_workbook(args.output_dir, date_value, header, data_rows)
        print(f"Saved: {saved_path} ({date_value.strftime('%Y-%m')})")
        time.sleep(0.2)


if __name__ == "__main__":
    main()
